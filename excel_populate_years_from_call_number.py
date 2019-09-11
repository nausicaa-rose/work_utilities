import re

import openpyxl

wb = openpyxl.load_workbook("Lendingrequestprintperiodicals-edited.xlsx")
sh = wb["2016-2019"]
year_p = re.compile(r" [12][890]\d\d$")

for c in sh.iter_rows(
    min_row=2, max_row=sh.max_row, min_col=24, max_col=24, values_only=True
):
    m = year_p.match(c[0].strip())
    if m is not None:
        years.append(m.groups()[0])
    else:
        years.append("")

year_col = sh.iter_rows(min_row=2, max_row=sh.max_row, min_col=12, max_col=12)

for v in years:
    c = next(year_col)
    c[0].value = v

wb.save("Lendingrequestprintperiodicals-edited2.xlsx")
